#!/usr/bin/env python3
"""REQ-105 AC4 / alpha-code#1108 —— 先证明判据能测出**已知的坏**,再拿它判未知的好。

本仓判据:「一个错误实现能不能满足这条断言?」本脚本把十种错误实现造出来,要求
read-xlsx-independent.py 对每一种都**红**,并且对唯一一份正确样本**绿**。
任何一条不符合预期,本次测量作废(退出码 1)。

依赖 openpyxl 只是为了**造样本**;判据本身(read-xlsx-independent.py)不碰 openpyxl。

用法:uv run --no-project --with openpyxl==3.1.5 selftest-known-bad.py \
        --contract fixture/xlsx-contract.json --workdir <tmp> [--json-out results/selftest.json]
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
READER = os.path.join(HERE, "read-xlsx-independent.py")


def build(path: str, sheets: list[tuple[str, dict]], keep_default: bool = False) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    for name, cells in sheets:
        ws = wb.create_sheet(name)
        for ref, value in cells.items():
            ws[ref] = value
    if not keep_default and "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    wb.save(path)


def run_reader(xlsx: str, contract_path: str) -> tuple[int, dict | None]:
    proc = subprocess.run(
        [sys.executable, READER, "--xlsx", xlsx, "--contract", contract_path],
        capture_output=True,
        text=True,
    )
    try:
        payload = json.loads(proc.stdout)
    except json.JSONDecodeError:
        payload = None
    return proc.returncode, payload


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--contract", required=True)
    ap.add_argument("--workdir")
    ap.add_argument("--json-out")
    args = ap.parse_args()

    with open(args.contract, encoding="utf-8") as fh:
        contract = json.load(fh)
    good = [(s["name"], dict(s["cells"])) for s in contract["sheets"]]

    workdir = args.workdir or tempfile.mkdtemp(prefix="req105-ac4-selftest-")
    os.makedirs(workdir, exist_ok=True)

    samples: list[tuple[str, str, bool]] = []  # (id, path, expect_green)

    def sample(sid: str, expect_green: bool) -> str:
        p = os.path.join(workdir, f"{sid}.xlsx")
        samples.append((sid, p, expect_green))
        return p

    # 0) the one sample that must be GREEN
    build(sample("correct", True), good)

    # 1) zero-byte file — "文件存在" 这条粗断言会被它满足
    open(sample("empty-file", False), "wb").close()

    # 2) not a zip at all
    with open(sample("not-a-zip", False), "w", encoding="utf-8") as fh:
        fh.write("this is not an OOXML package\n")

    # 3) 合法 xlsx,但一个单元格都没写(经典「写空文件的错误实现」)
    build(sample("empty-workbook", False), [("AC4Ledger", {}), ("AC4Structure", {})])

    # 4) 单元格值写错
    bad = [(n, dict(c)) for n, c in good]
    bad[0][1]["B2"] = 9999
    build(sample("wrong-cell-value", False), bad)

    # 5) 数字被写成文本 —— 值 "看起来" 对,类型错
    bad = [(n, dict(c)) for n, c in good]
    bad[0][1]["B2"] = "1108"
    build(sample("number-as-text", False), bad)

    # 6) 布尔被写成 1
    bad = [(n, dict(c)) for n, c in good]
    bad[0][1]["B3"] = 1
    build(sample("bool-as-number", False), bad)

    # 7) 只有一个 sheet(结构缺失)
    build(sample("missing-second-sheet", False), [good[0]])

    # 8) 默认 Sheet 没被摘掉
    build(sample("default-sheet-left", False), good, keep_default=True)

    # 9) sheet 顺序颠倒
    build(sample("sheet-order-swapped", False), [good[1], good[0]])

    # 10) 值对但落在别的 sheet 名下
    build(sample("renamed-sheet", False), [("Ledger", dict(good[0][1])), good[1]])

    rows = []
    ok = True
    for sid, path, expect_green in samples:
        code, payload = run_reader(path, args.contract)
        actually_green = code == 0
        red_ids = [c["id"] for c in (payload or {}).get("checks", []) if not c["ok"]]
        as_expected = actually_green == expect_green
        ok = ok and as_expected
        rows.append(
            {
                "sample": sid,
                "expected": "GREEN" if expect_green else "RED",
                "observed": "GREEN" if actually_green else "RED",
                "asExpected": as_expected,
                "redChecks": red_ids,
            }
        )

    out = {"workdir": workdir, "ok": ok, "samples": rows}
    payload = json.dumps(out, ensure_ascii=False, indent=2)
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as fh:
            fh.write(payload + "\n")
    print(payload)
    if not ok:
        print("本次测量作废:判据没能按预期区分已知的好与已知的坏", file=sys.stderr)
        return 1
    print(f"判据自检通过:{len(rows)} 个样本全部按预期(1 绿 / {len(rows) - 1} 红)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
